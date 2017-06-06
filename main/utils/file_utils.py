from os import listdir
from os.path import join, isfile, exists
import pandas as pd
from openpyxl import load_workbook, Workbook


def get_files(src_dir_path):
    return [f for f in listdir(src_dir_path) if isfile(join(src_dir_path, f))]


def read_csv_file(src_dir_path, f):
    data = pd.read_csv(join(src_dir_path, f), index_col=0, header=1)
    data.dropna(how='all')
    data.fillna(0, inplace=True)
    data.replace('(%)', '', inplace=True, regex=True)
    return data


def read_excel(src_dir_path, f, sheets_list):
    return pd.read_excel(join(src_dir_path, f), sheetname=sheets_list, header=0,index_col=0)


def write_excel(out_dir_path, df, sheet_name):
    book = Workbook()
    if not exists(out_dir_path):
        book.save(out_dir_path)
    book = load_workbook(out_dir_path)
    # book.remove_sheet(book.worksheets[0])
    writer = pd.ExcelWriter(out_dir_path, engine='openpyxl')
    writer.book = book
    for i in book.worksheets:
        if i.title == sheet_name:
            book.remove_sheet(i)
    df.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    return


def write_csv_file(path, data):
    data.to_csv(path, index=False)
    return
